#!/usr/bin/env python3
"""
Check for data continuity issues in Active Alproean List
"""

import openpyxl

def check_continuity():
    print("=" * 80)
    print("CHECKING DATA CONTINUITY IN ACTIVE ALPROEAN LIST")
    print("=" * 80)
    
    wb = openpyxl.load_workbook('/home/user/uploaded_files/Active Alproean List.xlsx')
    ws = wb.active
    
    print(f"\n📊 Total rows: {ws.max_row}")
    
    # Check for duplicate header rows or section breaks
    print("\n🔍 Checking for duplicate headers or section breaks...")
    
    header_rows = []
    empty_rows = []
    
    for row_idx in range(1, min(ws.max_row + 1, 100)):  # Check first 100 rows
        col_c = ws.cell(row=row_idx, column=3).value
        col_d = ws.cell(row=row_idx, column=4).value
        col_g = ws.cell(row=row_idx, column=7).value
        
        # Check if this looks like a header row
        if col_c and col_d:
            col_c_str = str(col_c).lower()
            col_d_str = str(col_d).lower()
            
            if 'nama' in col_c_str or 'name' in col_c_str or 'employee' in col_d_str:
                header_rows.append(row_idx)
                print(f"   ✅ Header row found at row {row_idx}: {col_c} | {col_d} | {col_g}")
        
        # Check if row is empty
        if not col_c and not col_d and not col_g:
            empty_rows.append(row_idx)
            if len(empty_rows) <= 5:
                print(f"   ⚠️ Empty row at {row_idx}")
    
    print(f"\n📊 Found {len(header_rows)} header row(s)")
    print(f"📊 Found {len(empty_rows)} empty row(s) in first 100 rows")
    
    # Check specific employees around rows 30-40
    print("\n\n" + "=" * 80)
    print("DETAILED VIEW: Rows 29-42 (where screenshot employees are)")
    print("=" * 80)
    print(f"\n{'Excel Row':<12} {'Name (Col C)':<30} {'ID (Col D)':<15} {'Role (Col G)':<40}")
    print("-" * 120)
    
    for row_idx in range(29, 43):
        col_c = ws.cell(row=row_idx, column=3).value or ""
        col_d = ws.cell(row=row_idx, column=4).value or ""
        col_g = ws.cell(row=row_idx, column=7).value or ""
        
        col_c_str = str(col_c)[:28]
        col_d_str = str(col_d)[:13]
        col_g_str = str(col_g)[:38]
        
        print(f"{row_idx:<12} {col_c_str:<30} {col_d_str:<15} {col_g_str:<40}")
    
    # Check for merged cells
    print("\n\n🔍 Checking for merged cells...")
    merged_ranges = list(ws.merged_cells.ranges)
    if merged_ranges:
        print(f"   ⚠️ Found {len(merged_ranges)} merged cell ranges:")
        for merge in merged_ranges[:10]:  # Show first 10
            print(f"      {merge}")
    else:
        print("   ✅ No merged cells found")
    
    # Simulate JavaScript array reading
    print("\n\n" + "=" * 80)
    print("SIMULATING JAVASCRIPT ARRAY READING (0-indexed)")
    print("=" * 80)
    print(f"\n{'Array Index':<15} {'Index 2 (Col C)':<30} {'Index 3 (Col D)':<15} {'Index 6 (Col G)':<40}")
    print("-" * 120)
    
    # In JavaScript, after slice(1), array index 0 = Excel row 2
    for excel_row in range(29, 43):
        js_array_index = excel_row - 2  # Because we do jsonData.slice(1), and arrays are 0-indexed
        col_c = ws.cell(row=excel_row, column=3).value or ""
        col_d = ws.cell(row=excel_row, column=4).value or ""
        col_g = ws.cell(row=excel_row, column=7).value or ""
        
        col_c_str = str(col_c)[:28]
        col_d_str = str(col_d)[:13]
        col_g_str = str(col_g)[:38]
        
        print(f"{js_array_index:<15} {col_c_str:<30} {col_d_str:<15} {col_g_str:<40}")

if __name__ == '__main__':
    try:
        check_continuity()
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
