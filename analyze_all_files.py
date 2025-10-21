#!/usr/bin/env python3
import openpyxl

def analyze_file(filename, title):
    print(f"\n{'='*70}")
    print(f"{title}")
    print(f"{'='*70}")
    
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    
    print(f"\nSheet name: {sheet.title}")
    print(f"Dimensions: {sheet.dimensions}")
    
    print("\nHeaders (Row 1):")
    headers = list(sheet[1])
    for idx, cell in enumerate(headers[:20]):  # First 20 columns
        col_letter = openpyxl.utils.get_column_letter(idx + 1)
        print(f"  Column {col_letter} (index {idx}): '{cell.value}'")
    
    print("\nFirst 3 data rows:")
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        if i > 4:
            break
        print(f"Row {i}: {row[:10]}")  # First 10 values

# Analyze each file
analyze_file('Mapping_AM.xlsx', 'OUTLET MAPPING FILE')
analyze_file('Sales_GP.xlsx', 'SALES & GP DATA FILE')
analyze_file('Personal_Sales.xlsx', 'PERSONAL SALES FILE')
