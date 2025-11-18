#!/usr/bin/env python3
"""
Deep analysis of Excel structure to understand column layout
"""

import openpyxl
import pandas as pd

def analyze_excel_structure():
    print("=" * 80)
    print("DEEP ANALYSIS OF EXCEL FILE STRUCTURE")
    print("=" * 80)
    
    # Use openpyxl to read raw structure
    print("\n📂 Loading Active Alproean List with openpyxl...")
    wb = openpyxl.load_workbook('/home/user/uploaded_files/Active Alproean List.xlsx')
    ws = wb.active
    
    print(f"\n📊 Worksheet name: {ws.title}")
    print(f"📊 Max row: {ws.max_row}")
    print(f"📊 Max column: {ws.max_column}")
    
    # Print first 20 rows, showing columns A through AO (1-41)
    print("\n🔍 RAW EXCEL DATA (First 20 rows, selected columns):")
    print("=" * 120)
    
    # Helper to get column letter
    def col_letter(idx):
        letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
                   'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO']
        return letters[idx-1] if idx <= len(letters) else f"Col{idx}"
    
    # Show specific columns: C(3), D(4), G(7), AO(41)
    print(f"\n{'Row':<5} {'Col C':<30} {'Col D':<20} {'Col G':<40} {'Col AO':<40}")
    print("-" * 140)
    
    for row_idx in range(1, min(21, ws.max_row + 1)):
        col_c = ws.cell(row=row_idx, column=3).value or ""  # Column C
        col_d = ws.cell(row=row_idx, column=4).value or ""  # Column D
        col_g = ws.cell(row=row_idx, column=7).value or ""  # Column G
        col_ao = ws.cell(row=row_idx, column=41).value or ""  # Column AO
        
        # Truncate long strings
        col_c_str = str(col_c)[:28]
        col_d_str = str(col_d)[:18]
        col_g_str = str(col_g)[:38]
        col_ao_str = str(col_ao)[:38]
        
        print(f"{row_idx:<5} {col_c_str:<30} {col_d_str:<20} {col_g_str:<40} {col_ao_str:<40}")
    
    # Find employees from screenshot
    print("\n\n" + "=" * 80)
    print("FINDING SPECIFIC EMPLOYEES FROM SCREENSHOT:")
    print("=" * 80)
    
    target_names = ['KHOO ZI YU', 'ESTER DESINDO NABABAN', 'RONTA SIREGAR', 
                    'LAELA FITIRAH', 'EPI PURNAMASARI', 'Mardian Rahayu', 'PUTRI AMBAR LESTARI']
    
    for target in target_names:
        print(f"\n🔍 Searching for: {target}")
        found = False
        for row_idx in range(1, min(ws.max_row + 1, 1000)):  # Search first 1000 rows
            col_c = ws.cell(row=row_idx, column=3).value
            if col_c and str(col_c).strip().upper() == target.upper():
                col_d = ws.cell(row=row_idx, column=4).value or ""
                col_g = ws.cell(row=row_idx, column=7).value or ""
                col_ao = ws.cell(row=row_idx, column=41).value or ""
                print(f"   ✅ Found at Row {row_idx}:")
                print(f"      Name (Col C): {col_c}")
                print(f"      ID (Col D): {col_d}")
                print(f"      Role (Col G): {col_g}")
                print(f"      Outlet (Col AO): {col_ao}")
                found = True
                break
        
        if not found:
            print(f"   ❌ Not found")
    
    # Now check what the JavaScript would read
    print("\n\n" + "=" * 80)
    print("SIMULATING JAVASCRIPT XLSX.read():")
    print("=" * 80)
    
    # Read with pandas as arrays (similar to what JavaScript does)
    df = pd.read_excel('/home/user/uploaded_files/Active Alproean List.xlsx', header=None)
    
    print(f"\n📊 Array-based reading (no headers):")
    print(f"   Shape: {df.shape}")
    print(f"\n   First 15 rows, columns C(2), D(3), G(6), AO(40) [0-indexed]:")
    print("-" * 120)
    print(f"{'Row':<5} {'Index 2 (Col C)':<30} {'Index 3 (Col D)':<20} {'Index 6 (Col G)':<40} {'Index 40 (Col AO)':<40}")
    print("-" * 120)
    
    for idx in range(min(15, len(df))):
        val_2 = str(df.iloc[idx, 2])[:28] if pd.notna(df.iloc[idx, 2]) else ""
        val_3 = str(df.iloc[idx, 3])[:18] if pd.notna(df.iloc[idx, 3]) else ""
        val_6 = str(df.iloc[idx, 6])[:38] if pd.notna(df.iloc[idx, 6]) else ""
        val_40 = str(df.iloc[idx, 40])[:38] if pd.notna(df.iloc[idx, 40]) else ""
        
        print(f"{idx:<5} {val_2:<30} {val_3:<20} {val_6:<40} {val_40:<40}")

if __name__ == '__main__':
    try:
        analyze_excel_structure()
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
