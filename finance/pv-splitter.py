#!/usr/bin/env python3
"""
PV Splitter Backend Service
Splits Excel files with multiple sheets into individual files
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import openpyxl
from openpyxl.utils import get_column_letter
import io
import zipfile
import os
import json
from datetime import datetime
import tempfile
import shutil
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

app = Flask(__name__)
CORS(app)

# Email Configuration
# These should be set via environment variables or config file
EMAIL_CONFIG = {
    'smtp_server': os.getenv('SMTP_SERVER', 'smtp.gmail.com'),
    'smtp_port': int(os.getenv('SMTP_PORT', '587')),
    'sender_email': os.getenv('SENDER_EMAIL', ''),
    'sender_password': os.getenv('SENDER_PASSWORD', ''),
    'sender_name': os.getenv('SENDER_NAME', 'Apotek Alpro Finance Team')
}

@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'service': 'pv-splitter',
        'timestamp': datetime.now().isoformat()
    })

@app.route('/api/split-excel', methods=['POST'])
def split_excel():
    """
    Split an Excel file with multiple sheets into individual files
    Returns metadata about the split files for download
    """
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'File must be an Excel file (.xlsx or .xls)'}), 400
        
        # Read the uploaded Excel file
        file_content = file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(file_content))
        
        # Create a temporary directory to store split files
        temp_dir = tempfile.mkdtemp()
        split_files_info = []
        
        # Process each sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Extract company name from D9
            company_name = sheet['D9'].value
            if company_name is None:
                company_name = f"Sheet_{sheet_name}"
            else:
                company_name = str(company_name).strip()
            
            # Extract PV number from R9
            pv_number = sheet['R9'].value
            if pv_number is None:
                pv_number = "NO_PV"
            else:
                pv_number = str(pv_number).strip()
            
            # Extract email from D12
            email = sheet['D12'].value
            if email is None:
                email = ""
            else:
                email = str(email).strip()
            
            # Create safe filename
            # Replace invalid filename characters (including /)
            safe_company = "".join(c for c in company_name if c.isalnum() or c in (' ', '-', '_')).strip()
            safe_pv = "".join(c if (c.isalnum() or c in (' ', '-', '_')) else '-' for c in pv_number).strip()
            
            # Create filename: "CompanyName PV_Number.xlsx"
            filename = f"{safe_company} {safe_pv}.xlsx"
            
            # Create a new workbook with only this sheet
            new_workbook = openpyxl.Workbook()
            new_workbook.remove(new_workbook.active)  # Remove default sheet
            
            # Copy the sheet to the new workbook
            new_sheet = new_workbook.create_sheet(sheet_name)
            
            # Copy all cells with their values, styles, and formatting
            for row in sheet.iter_rows():
                for cell in row:
                    new_cell = new_sheet[cell.coordinate]
                    new_cell.value = cell.value
                    
                    # Copy cell formatting
                    if cell.has_style:
                        new_cell.font = cell.font.copy()
                        new_cell.border = cell.border.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.number_format = cell.number_format
                        new_cell.protection = cell.protection.copy()
                        new_cell.alignment = cell.alignment.copy()
            
            # Copy column dimensions
            for col in sheet.column_dimensions:
                if col in sheet.column_dimensions:
                    new_sheet.column_dimensions[col].width = sheet.column_dimensions[col].width
            
            # Copy row dimensions
            for row in sheet.row_dimensions:
                if row in sheet.row_dimensions:
                    new_sheet.row_dimensions[row].height = sheet.row_dimensions[row].height
            
            # Copy merged cells
            for merged_cell in sheet.merged_cells.ranges:
                new_sheet.merge_cells(str(merged_cell))
            
            # Save the new workbook
            file_path = os.path.join(temp_dir, filename)
            new_workbook.save(file_path)
            
            # Store file information
            split_files_info.append({
                'filename': filename,
                'company_name': company_name,
                'pv_number': pv_number,
                'email': email,
                'sheet_name': sheet_name,
                'file_path': file_path
            })
        
        # Create a zip file containing all split files
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for file_info in split_files_info:
                zip_file.write(file_info['file_path'], file_info['filename'])
        
        zip_buffer.seek(0)
        
        # Store zip file temporarily
        zip_filename = f"pv_split_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        zip_path = os.path.join(temp_dir, zip_filename)
        with open(zip_path, 'wb') as f:
            f.write(zip_buffer.getvalue())
        
        # Return file information (without file_path for security)
        response_data = {
            'success': True,
            'files': [{
                'filename': f['filename'],
                'company_name': f['company_name'],
                'pv_number': f['pv_number'],
                'email': f['email'],
                'sheet_name': f['sheet_name']
            } for f in split_files_info],
            'total_files': len(split_files_info),
            'zip_filename': zip_filename,
            'temp_dir': temp_dir  # Full path for backend use
        }
        
        return jsonify(response_data)
    
    except Exception as e:
        return jsonify({
            'error': str(e),
            'type': type(e).__name__
        }), 500

@app.route('/api/download-zip/<path:temp_dir>/<filename>', methods=['GET'])
def download_zip(temp_dir, filename):
    """Download the zip file containing all split files"""
    try:
        # Reconstruct the full path - temp_dir comes from URL without leading /
        if not temp_dir.startswith('/'):
            temp_dir = '/' + temp_dir
        file_path = os.path.join(temp_dir, filename)
        
        if not os.path.exists(file_path):
            return jsonify({'error': f'File not found: {file_path}'}), 404
        
        def cleanup():
            """Cleanup temporary directory after download"""
            try:
                if os.path.exists(temp_dir):
                    shutil.rmtree(temp_dir)
            except:
                pass
        
        response = send_file(
            file_path,
            mimetype='application/zip',
            as_attachment=True,
            download_name=filename
        )
        
        # Schedule cleanup (note: this won't work perfectly in all cases)
        # For production, consider a background cleanup task
        
        return response
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/download-individual/<path:temp_dir>/<filename>', methods=['GET'])
def download_individual(temp_dir, filename):
    """Download an individual split file"""
    try:
        # Reconstruct the full path - temp_dir comes from URL without leading /
        if not temp_dir.startswith('/'):
            temp_dir = '/' + temp_dir
        file_path = os.path.join(temp_dir, filename)
        
        if not os.path.exists(file_path):
            return jsonify({'error': f'File not found: {file_path}'}), 404
        
        return send_file(
            file_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/get-email-data', methods=['POST'])
def get_email_data():
    """
    Prepare email data for Gmail integration
    Returns the email recipients and their corresponding files
    """
    try:
        data = request.json
        files_info = data.get('files', [])
        
        # Group files by email
        email_groups = {}
        for file_info in files_info:
            email = file_info.get('email', '').strip()
            if email and '@' in email:  # Basic email validation
                if email not in email_groups:
                    email_groups[email] = []
                email_groups[email].append(file_info)
        
        # Prepare email data
        email_data = []
        for email, files in email_groups.items():
            email_data.append({
                'to': email,
                'subject': f"PV Documents - {', '.join([f['pv_number'] for f in files[:3]])}{'...' if len(files) > 3 else ''}",
                'body': f"""Dear {files[0]['company_name']},

Please find attached your PV document(s):

{chr(10).join([f"- {f['filename']}" for f in files])}

Best regards,
Apotek Alpro Finance Team""",
                'files': files
            })
        
        return jsonify({
            'success': True,
            'email_count': len(email_data),
            'emails': email_data
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/send-emails', methods=['POST'])
def send_emails():
    """
    Send emails with file attachments automatically
    Requires email configuration to be set
    """
    try:
        data = request.json
        emails_to_send = data.get('emails', [])
        temp_dir = data.get('temp_dir', '')
        
        # Check if email is configured
        if not EMAIL_CONFIG['sender_email'] or not EMAIL_CONFIG['sender_password']:
            return jsonify({
                'error': 'Email not configured. Please set SENDER_EMAIL and SENDER_PASSWORD environment variables.',
                'configured': False
            }), 400
        
        results = []
        errors = []
        
        for email_data in emails_to_send:
            try:
                # Send individual email
                success = send_single_email(
                    to_email=email_data['to'],
                    subject=email_data['subject'],
                    body=email_data['body'],
                    files=email_data['files'],
                    temp_dir=temp_dir
                )
                
                if success:
                    results.append({
                        'to': email_data['to'],
                        'status': 'sent',
                        'files_count': len(email_data['files'])
                    })
                else:
                    errors.append({
                        'to': email_data['to'],
                        'error': 'Failed to send email'
                    })
                    
            except Exception as e:
                errors.append({
                    'to': email_data.get('to', 'unknown'),
                    'error': str(e)
                })
        
        return jsonify({
            'success': len(errors) == 0,
            'sent': len(results),
            'failed': len(errors),
            'results': results,
            'errors': errors
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def send_single_email(to_email, subject, body, files, temp_dir):
    """
    Send a single email with attachments using SMTP
    """
    try:
        # Create message
        msg = MIMEMultipart()
        msg['From'] = f"{EMAIL_CONFIG['sender_name']} <{EMAIL_CONFIG['sender_email']}>"
        msg['To'] = to_email
        msg['Subject'] = subject
        
        # Add body
        msg.attach(MIMEText(body, 'plain'))
        
        # Reconstruct temp_dir path if needed
        if not temp_dir.startswith('/'):
            temp_dir = '/' + temp_dir
        
        # Attach files
        for file_info in files:
            filename = file_info['filename']
            file_path = os.path.join(temp_dir, filename)
            
            if os.path.exists(file_path):
                with open(file_path, 'rb') as f:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(f.read())
                    encoders.encode_base64(part)
                    part.add_header(
                        'Content-Disposition',
                        f'attachment; filename= {filename}'
                    )
                    msg.attach(part)
        
        # Connect to SMTP server and send
        with smtplib.SMTP(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port']) as server:
            server.starttls()
            server.login(EMAIL_CONFIG['sender_email'], EMAIL_CONFIG['sender_password'])
            server.send_message(msg)
        
        return True
        
    except Exception as e:
        print(f"Error sending email to {to_email}: {str(e)}")
        return False

@app.route('/api/test-email-config', methods=['GET'])
def test_email_config():
    """
    Test if email configuration is set up correctly
    """
    config_status = {
        'configured': bool(EMAIL_CONFIG['sender_email'] and EMAIL_CONFIG['sender_password']),
        'smtp_server': EMAIL_CONFIG['smtp_server'],
        'smtp_port': EMAIL_CONFIG['smtp_port'],
        'sender_email': EMAIL_CONFIG['sender_email'][:3] + '***' if EMAIL_CONFIG['sender_email'] else 'Not set',
        'sender_name': EMAIL_CONFIG['sender_name'],
        'password_set': bool(EMAIL_CONFIG['sender_password'])
    }
    
    return jsonify(config_status)

if __name__ == '__main__':
    print("Starting PV Splitter Service on port 5001...")
    print("Service endpoints:")
    print("  - POST /api/split-excel - Split Excel file")
    print("  - GET  /api/download-zip/<temp_dir>/<filename> - Download all files as zip")
    print("  - GET  /api/download-individual/<temp_dir>/<filename> - Download individual file")
    print("  - POST /api/get-email-data - Get email data for Gmail integration")
    print("  - POST /api/send-emails - Send emails automatically with attachments")
    print("  - GET  /api/test-email-config - Test email configuration")
    print("  - GET  /health - Health check")
    print("")
    print("Email Configuration:")
    print(f"  - SMTP Server: {EMAIL_CONFIG['smtp_server']}:{EMAIL_CONFIG['smtp_port']}")
    print(f"  - Sender Email: {EMAIL_CONFIG['sender_email'] if EMAIL_CONFIG['sender_email'] else 'NOT CONFIGURED'}")
    print(f"  - Status: {'READY' if EMAIL_CONFIG['sender_email'] and EMAIL_CONFIG['sender_password'] else 'NOT CONFIGURED - Set SENDER_EMAIL and SENDER_PASSWORD'}")
    print("")
    app.run(host='0.0.0.0', port=5001, debug=True)
