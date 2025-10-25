#!/usr/bin/env python3
"""
Serah Terima IM Split Backend API
Handles Excel file processing, column deletion, supplier-based splitting,
and WhatsApp link generation for file distribution.
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import openpyxl
from openpyxl.utils import get_column_letter
import os
import tempfile
import zipfile
from io import BytesIO
import re
from datetime import datetime
import urllib.parse
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

app = Flask(__name__)
CORS(app)

# Global storage for temp directories
temp_dirs = {}

# Email Configuration
EMAIL_CONFIG = {
    'smtp_server': os.getenv('SMTP_SERVER', 'smtp.gmail.com'),
    'smtp_port': int(os.getenv('SMTP_PORT', '587')),
    'sender_email': os.getenv('SENDER_EMAIL', ''),
    'sender_password': os.getenv('SENDER_PASSWORD', ''),
    'sender_name': os.getenv('SENDER_NAME', 'Apotek Alpro Finance Team')
}


def extract_supplier_name(full_supplier_text):
    """
    Extract supplier name from format: '0000000008 - PT. INDOCORE PERKASA'
    - Take first 13 characters
    - Remove numerics and dash
    - Return only alphabetic part
    
    Example: '0000000008 - PT. INDOCORE PERKASA' -> 'PT. INDOCORE PERKASA'
    """
    if not full_supplier_text:
        return "UNKNOWN"
    
    # Take first 13 characters
    first_13 = str(full_supplier_text)[:13]
    
    # Remove numbers and dashes, keep only letters and spaces
    cleaned = re.sub(r'[0-9\-]', '', first_13).strip()
    
    # If still empty after cleaning, use the full text after dash
    if not cleaned and ' - ' in str(full_supplier_text):
        cleaned = str(full_supplier_text).split(' - ', 1)[1].strip()
    
    return cleaned if cleaned else "UNKNOWN"


def extract_date_from_cell(cell_value):
    """
    Extract date from cell G3 format: '2025-10-24 18:13:33'
    Return format: '20251024'
    """
    if not cell_value:
        return datetime.now().strftime('%Y%m%d')
    
    # Convert to string
    date_str = str(cell_value)
    
    # Try to extract date pattern YYYY-MM-DD
    match = re.search(r'(\d{4})-(\d{2})-(\d{2})', date_str)
    if match:
        return match.group(1) + match.group(2) + match.group(3)
    
    # Fallback to current date
    return datetime.now().strftime('%Y%m%d')


def copy_cell_with_style(source_cell, target_cell):
    """Copy cell value and style from source to target."""
    target_cell.value = source_cell.value
    
    if source_cell.has_style:
        target_cell.font = source_cell.font.copy()
        target_cell.border = source_cell.border.copy()
        target_cell.fill = source_cell.fill.copy()
        target_cell.number_format = source_cell.number_format
        target_cell.protection = source_cell.protection.copy()
        target_cell.alignment = source_cell.alignment.copy()


def delete_columns_and_split(file_path, wa_mapping, email_mapping=None):
    """
    Main processing function:
    1. Read original file
    2. Extract date from G3 (before deletion)
    3. Delete columns B, C, E, F, H, I, M, P, S, T, U
    4. Group by supplier (original column D, becomes column B)
    5. Create separate files per supplier
    6. Generate WhatsApp links
    7. Generate Email mappings
    """
    if email_mapping is None:
        email_mapping = {}
    
    # Columns to delete (by original position, 1-indexed)
    # B=2, C=3, E=5, F=6, H=8, I=9, M=13, P=16, S=19, T=20, U=21
    columns_to_delete = [21, 20, 19, 16, 13, 9, 8, 6, 5, 3, 2]  # Reverse order for deletion
    
    # Load workbook
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # Extract date from G3 BEFORE deletion
    date_str = extract_date_from_cell(sheet['G3'].value)
    
    # Get all rows as list
    all_rows = list(sheet.iter_rows())
    
    # Close original workbook
    wb.close()
    
    # Group data by supplier
    # First, identify header row and data rows
    header_row_idx = None
    for idx, row in enumerate(all_rows):
        if row[3].value == 'Supplier':  # Column D (index 3) = Supplier header
            header_row_idx = idx
            break
    
    if header_row_idx is None:
        raise ValueError("Could not find 'Supplier' header row")
    
    # Group rows by supplier
    supplier_groups = {}
    header_rows = all_rows[:header_row_idx + 1]  # Include all rows up to and including header
    data_rows = all_rows[header_row_idx + 1:]  # Data rows after header
    
    for row in data_rows:
        supplier = row[3].value  # Column D (index 3)
        if supplier and str(supplier).strip():
            if supplier not in supplier_groups:
                supplier_groups[supplier] = []
            supplier_groups[supplier].append(row)
    
    # Create temp directory for split files
    temp_dir = tempfile.mkdtemp()
    split_files = []
    
    # Process each supplier group
    for supplier, rows in supplier_groups.items():
        # Create new workbook for this supplier
        new_wb = openpyxl.Workbook()
        new_wb.remove(new_wb.active)  # Remove default sheet
        new_sheet = new_wb.create_sheet(sheet.title)
        
        # Copy header rows
        for src_row_idx, src_row in enumerate(header_rows, 1):
            for src_col_idx, src_cell in enumerate(src_row, 1):
                # Skip columns that will be deleted
                if src_col_idx in [2, 3, 5, 6, 8, 9, 13, 16, 19, 20, 21]:
                    continue
                
                # Calculate target column index (after deletions)
                target_col_idx = src_col_idx
                for del_col in [2, 3, 5, 6, 8, 9, 13, 16, 19, 20, 21]:
                    if src_col_idx > del_col:
                        target_col_idx -= 1
                
                target_cell = new_sheet.cell(row=src_row_idx, column=target_col_idx)
                copy_cell_with_style(src_cell, target_cell)
        
        # Copy data rows for this supplier
        target_row_idx = len(header_rows) + 1
        for src_row in rows:
            for src_col_idx, src_cell in enumerate(src_row, 1):
                # Skip columns that will be deleted
                if src_col_idx in [2, 3, 5, 6, 8, 9, 13, 16, 19, 20, 21]:
                    continue
                
                # Calculate target column index
                target_col_idx = src_col_idx
                for del_col in [2, 3, 5, 6, 8, 9, 13, 16, 19, 20, 21]:
                    if src_col_idx > del_col:
                        target_col_idx -= 1
                
                target_cell = new_sheet.cell(row=target_row_idx, column=target_col_idx)
                copy_cell_with_style(src_cell, target_cell)
            
            target_row_idx += 1
        
        # Auto-fit all columns to content for professional look
        for column in new_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Set column width with some padding
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            if adjusted_width > 0:
                new_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Generate filename
        supplier_name = extract_supplier_name(supplier)
        safe_supplier_name = "".join(c for c in supplier_name if c.isalnum() or c in (' ', '-', '_')).strip()
        safe_supplier_name = safe_supplier_name.replace(' ', '_')
        filename = f"{safe_supplier_name}_{date_str}.xlsx"
        filepath = os.path.join(temp_dir, filename)
        
        # Save file
        new_wb.save(filepath)
        new_wb.close()
        
        # Get WhatsApp group/contact from mapping
        # Try multiple matching strategies
        wa_target = None
        
        # Strategy 1: Exact match with full supplier string
        wa_target = wa_mapping.get(supplier, None)
        
        # Strategy 2: Try matching just the company name part (after dash)
        if not wa_target and ' - ' in str(supplier):
            company_only = str(supplier).split(' - ', 1)[1].strip()
            wa_target = wa_mapping.get(company_only, None)
        
        # Strategy 3: Try matching with cleaned supplier name
        if not wa_target:
            wa_target = wa_mapping.get(supplier_name, None)
        
        # Get Email from mapping (same matching strategies)
        email_address = None
        
        # Strategy 1: Exact match with full supplier string
        email_address = email_mapping.get(supplier, None)
        
        # Strategy 2: Try matching just the company name part (after dash)
        if not email_address and ' - ' in str(supplier):
            company_only = str(supplier).split(' - ', 1)[1].strip()
            email_address = email_mapping.get(company_only, None)
        
        # Strategy 3: Try matching with cleaned supplier name
        if not email_address:
            email_address = email_mapping.get(supplier_name, None)
        
        split_files.append({
            'filename': filename,
            'supplier': supplier,
            'supplier_clean': supplier_name,
            'row_count': len(rows),
            'wa_target': wa_target,
            'email': email_address,
            'filepath': filepath
        })
    
    return temp_dir, split_files, date_str


@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint."""
    return jsonify({'status': 'healthy', 'service': 'IM Splitter API'}), 200


@app.route('/api/split-im-excel', methods=['POST'])
def split_im_excel():
    """
    Process uploaded Excel file:
    1. Delete specified columns
    2. Split by supplier
    3. Generate WhatsApp links (optional)
    4. Generate email mappings (optional)
    """
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Save uploaded files temporarily
        temp_input = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        file.save(temp_input.name)
        temp_input.close()
        
        # Read WhatsApp mapping (optional)
        wa_mapping = {}
        if 'wa_mapping' in request.files and request.files['wa_mapping'].filename != '':
            wa_file = request.files['wa_mapping']
            temp_wa = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            wa_file.save(temp_wa.name)
            temp_wa.close()
            
            try:
                wb_wa = openpyxl.load_workbook(temp_wa.name)
                sheet_wa = wb_wa.active
                
                for row in sheet_wa.iter_rows(min_row=2, values_only=True):  # Skip header
                    if row[0] and row[1]:  # Column A = Supplier, Column B = WA Group
                        supplier_name = str(row[0]).strip()
                        wa_group = str(row[1]).strip()
                        
                        # Store both exact match and just the company name part
                        wa_mapping[supplier_name] = wa_group
                        
                        # Also create a mapping for full format with numbers
                        # In case the mapping file has format: 0000000008 - PT. INDOCORE PERKASA
                        if ' - ' in supplier_name:
                            company_only = supplier_name.split(' - ', 1)[1].strip()
                            wa_mapping[company_only] = wa_group
                
                wb_wa.close()
            except Exception as e:
                return jsonify({"error": f"Error reading WhatsApp mapping: {str(e)}"}), 400
            finally:
                os.unlink(temp_wa.name)
        
        # Read Email mapping (optional)
        email_mapping = {}
        if 'email_mapping' in request.files and request.files['email_mapping'].filename != '':
            email_file = request.files['email_mapping']
            temp_email = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            email_file.save(temp_email.name)
            temp_email.close()
            
            try:
                wb_email = openpyxl.load_workbook(temp_email.name)
                sheet_email = wb_email.active
                
                for row in sheet_email.iter_rows(min_row=2, values_only=True):  # Skip header
                    if row[0] and row[1]:  # Column A = Supplier, Column B = Email
                        supplier_name = str(row[0]).strip()
                        email_addr = str(row[1]).strip()
                        
                        # Store both exact match and just the company name part
                        email_mapping[supplier_name] = email_addr
                        
                        # Also create a mapping for full format with numbers
                        if ' - ' in supplier_name:
                            company_only = supplier_name.split(' - ', 1)[1].strip()
                            email_mapping[company_only] = email_addr
                
                wb_email.close()
            except Exception as e:
                return jsonify({"error": f"Error reading Email mapping: {str(e)}"}), 400
            finally:
                os.unlink(temp_email.name)
        
        # Process the file
        temp_dir, split_files, date_str = delete_columns_and_split(temp_input.name, wa_mapping, email_mapping)
        
        # Clean up input file
        os.unlink(temp_input.name)
        
        # Store temp directory reference
        temp_dirs[temp_dir] = True
        
        # Create ZIP file
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for file_info in split_files:
                zip_file.write(file_info['filepath'], file_info['filename'])
        
        zip_buffer.seek(0)
        zip_filename = f'IM_Split_{date_str}.zip'
        zip_path = os.path.join(temp_dir, zip_filename)
        
        with open(zip_path, 'wb') as f:
            f.write(zip_buffer.getvalue())
        
        # Prepare response
        result = {
            'success': True,
            'date': date_str,
            'temp_dir': temp_dir,
            'zip_filename': zip_filename,
            'files': split_files,
            'total_suppliers': len(split_files)
        }
        
        return jsonify(result), 200
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error in split_im_excel: {error_trace}")
        return jsonify({'error': str(e), 'trace': error_trace}), 500


@app.route('/api/download-zip/<path:temp_dir>/<filename>', methods=['GET'])
def download_zip(temp_dir, filename):
    """Download the ZIP file containing all split files."""
    try:
        # Reconstruct full path
        if not temp_dir.startswith('/'):
            temp_dir = '/' + temp_dir
        
        filepath = os.path.join(temp_dir, filename)
        
        if not os.path.exists(filepath):
            return jsonify({'error': 'File not found'}), 404
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/zip'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/download-file/<path:temp_dir>/<filename>', methods=['GET'])
def download_file(temp_dir, filename):
    """Download individual split file."""
    try:
        # Reconstruct full path
        if not temp_dir.startswith('/'):
            temp_dir = '/' + temp_dir
        
        filepath = os.path.join(temp_dir, filename)
        
        if not os.path.exists(filepath):
            return jsonify({'error': 'File not found'}), 404
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/generate-wa-link/<path:temp_dir>/<filename>', methods=['POST'])
def generate_wa_link(temp_dir, filename):
    """
    Generate WhatsApp Web link with text message.
    File must be downloaded separately and attached manually.
    """
    try:
        data = request.json
        wa_target = data.get('wa_target', '')
        supplier_name = data.get('supplier_name', '')
        
        # Create message text
        message = f"Serah Terima IM - {supplier_name}\n\nFile: {filename}\n\nMohon di-download dan di-check. Terima kasih!"
        
        # Encode message for URL
        encoded_message = urllib.parse.quote(message)
        
        # WhatsApp Web link (works for both personal and group)
        # Note: Group invites need special link format, but we'll use simple text link
        wa_link = f"https://web.whatsapp.com/send?text={encoded_message}"
        
        # For phone numbers, use: https://wa.me/{phone}?text={message}
        # For groups, user needs to paste in the group chat
        
        result = {
            'success': True,
            'wa_link': wa_link,
            'message': message,
            'note': 'Please download the file first, then click this link to open WhatsApp and attach the file manually.'
        }
        
        return jsonify(result), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/send-emails', methods=['POST'])
def send_emails():
    """
    Send emails with file attachments automatically
    Expects: {emails: [{to, subject, body, files}], temp_dir}
    """
    try:
        data = request.json
        emails_to_send = data.get('emails', [])
        temp_dir = data.get('temp_dir', '')
        
        # Check if email is configured
        if not EMAIL_CONFIG['sender_email'] or not EMAIL_CONFIG['sender_password']:
            return jsonify({
                'error': 'Email not configured. Please set SENDER_EMAIL and SENDER_PASSWORD environment variables.',
                'configured': False
            }), 400
        
        results = []
        errors = []
        
        for email_data in emails_to_send:
            try:
                # Send individual email
                success = send_single_email(
                    to_email=email_data['to'],
                    subject=email_data['subject'],
                    body=email_data['body'],
                    files=email_data.get('files', []),
                    temp_dir=temp_dir
                )
                
                if success:
                    results.append({
                        'to': email_data['to'],
                        'status': 'sent',
                        'files_count': len(email_data.get('files', []))
                    })
                else:
                    errors.append({
                        'to': email_data['to'],
                        'error': 'Failed to send email'
                    })
                    
            except Exception as e:
                errors.append({
                    'to': email_data.get('to', 'unknown'),
                    'error': str(e)
                })
        
        return jsonify({
            'success': len(errors) == 0,
            'sent': len(results),
            'failed': len(errors),
            'results': results,
            'errors': errors
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def send_single_email(to_email, subject, body, files, temp_dir):
    """
    Send a single email with attachments using SMTP
    """
    try:
        # Create message
        msg = MIMEMultipart()
        msg['From'] = f"{EMAIL_CONFIG['sender_name']} <{EMAIL_CONFIG['sender_email']}>"
        msg['To'] = to_email
        msg['Subject'] = subject
        
        # Add body
        msg.attach(MIMEText(body, 'plain'))
        
        # Reconstruct temp_dir path if needed
        if not temp_dir.startswith('/'):
            temp_dir = '/' + temp_dir
        
        # Attach files
        for file_info in files:
            filename = file_info.get('filename', '')
            file_path = os.path.join(temp_dir, filename)
            
            if os.path.exists(file_path):
                with open(file_path, 'rb') as f:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(f.read())
                    encoders.encode_base64(part)
                    part.add_header(
                        'Content-Disposition',
                        f'attachment; filename= {filename}'
                    )
                    msg.attach(part)
        
        # Connect to SMTP server and send
        with smtplib.SMTP(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port']) as server:
            server.starttls()
            server.login(EMAIL_CONFIG['sender_email'], EMAIL_CONFIG['sender_password'])
            server.send_message(msg)
        
        return True
        
    except Exception as e:
        print(f"Error sending email to {to_email}: {str(e)}")
        return False


@app.route('/api/test-email-config', methods=['GET'])
def test_email_config():
    """
    Test if email configuration is set up correctly
    """
    config_status = {
        'configured': bool(EMAIL_CONFIG['sender_email'] and EMAIL_CONFIG['sender_password']),
        'smtp_server': EMAIL_CONFIG['smtp_server'],
        'smtp_port': EMAIL_CONFIG['smtp_port'],
        'sender_email': EMAIL_CONFIG['sender_email'][:3] + '***' if EMAIL_CONFIG['sender_email'] else 'Not set',
        'sender_name': EMAIL_CONFIG['sender_name'],
        'password_set': bool(EMAIL_CONFIG['sender_password'])
    }
    
    return jsonify(config_status)


if __name__ == '__main__':
    print("=" * 60)
    print("IM Splitter API Server Starting...")
    print("=" * 60)
    print("Endpoints:")
    print("  POST /api/split-im-excel - Process and split Excel file")
    print("  GET  /api/download-zip/<temp_dir>/<filename> - Download ZIP")
    print("  GET  /api/download-file/<temp_dir>/<filename> - Download single file")
    print("  POST /api/generate-wa-link/<temp_dir>/<filename> - Generate WhatsApp link")
    print("  POST /api/send-emails - Send emails with attachments")
    print("  GET  /api/test-email-config - Test email configuration")
    print("  GET  /health - Health check")
    print("=" * 60)
    print("Email Configuration:")
    print(f"  SMTP Server: {EMAIL_CONFIG['smtp_server']}:{EMAIL_CONFIG['smtp_port']}")
    print(f"  Sender Email: {EMAIL_CONFIG['sender_email'] if EMAIL_CONFIG['sender_email'] else 'NOT CONFIGURED'}")
    print(f"  Status: {'READY' if EMAIL_CONFIG['sender_email'] and EMAIL_CONFIG['sender_password'] else 'NOT CONFIGURED - Set SENDER_EMAIL and SENDER_PASSWORD'}")
    print("=" * 60)
    app.run(host='0.0.0.0', port=5002, debug=True)
